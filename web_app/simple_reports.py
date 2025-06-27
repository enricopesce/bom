"""
Simplified Report Generator for VM Assessment BOM Tool
Creates exactly 3 optimized output files: Excel, Text, and CSV
"""

import csv
import json
import re
from typing import List, Dict, Any
from datetime import datetime
from pathlib import Path
from io import StringIO

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Professional formatting libraries
try:
    from rich.console import Console
    from rich.table import Table
    from rich.panel import Panel
    from rich.columns import Columns
    from rich.text import Text
    from rich.align import Align
    from rich import box
    from tabulate import tabulate
    FORMATTING_AVAILABLE = True
except ImportError:
    FORMATTING_AVAILABLE = False

from models.vm_models import VMAssessment, BillOfMaterials, OSType


class SimplifiedReportGenerator:
    """Simplified report generator that creates exactly 3 files: Excel, Text, CSV"""
    
    def __init__(self, output_dir: str):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def generate_all_reports(self, bom: BillOfMaterials, assessment: VMAssessment, source_filename: str = None) -> Dict[str, str]:
        """
        Generate all three report formats
        
        Args:
            bom: Bill of Materials with cost calculations
            assessment: VM Assessment with VM details
            source_filename: Original filename for better naming
        
        Returns:
            Dict with file paths for each format
        """
        files = {}
        
        # Create base filename from source
        base_name = self._create_base_filename(source_filename)
        
        # Generate Excel report
        if EXCEL_AVAILABLE:
            excel_file = self.output_dir / f"{base_name}_BOM_Report.xlsx"
            self._generate_excel_report(bom, assessment, excel_file)
            files['excel'] = str(excel_file)
        
        # Generate Text report
        text_file = self.output_dir / f"{base_name}_BOM_Report.txt"
        self._generate_text_report(bom, assessment, text_file)
        files['text'] = str(text_file)
        
        # Generate CSV report
        csv_file = self.output_dir / f"{base_name}_BOM_Data.csv"
        self._generate_csv_report(bom, assessment, csv_file)
        files['csv'] = str(csv_file)
        
        return files
    
    def _create_base_filename(self, source_filename: str = None) -> str:
        """Create a clean base filename from the source filename"""
        if not source_filename:
            return f"VM_Assessment_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        # Extract filename without extension
        base = Path(source_filename).stem
        
        # Clean up the filename - remove common RVTools prefixes and normalize
        base = re.sub(r'^RVTools[_\s]*export[_\s]*', '', base, flags=re.IGNORECASE)
        base = re.sub(r'^RVTools[_\s]*', '', base, flags=re.IGNORECASE)
        base = re.sub(r'[_\s]*all[_\s]*', '_', base, flags=re.IGNORECASE)
        
        # Replace spaces and special characters with underscores
        base = re.sub(r'[^\w\-]', '_', base)
        
        # Remove multiple underscores
        base = re.sub(r'_+', '_', base)
        
        # Remove leading/trailing underscores
        base = base.strip('_')
        
        # Ensure it's not empty and not too long
        if not base or len(base) < 3:
            base = f"VM_Assessment_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        elif len(base) > 50:
            base = base[:50].rstrip('_')
        
        return base
    
    def _generate_excel_report(self, bom: BillOfMaterials, assessment: VMAssessment, output_file: Path):
        """Generate comprehensive Excel report with multiple sheets"""
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # 1. Summary Sheet
        self._create_summary_sheet(wb, bom, assessment)
        
        # 2. VM Details Sheet
        self._create_vm_details_sheet(wb, bom, assessment)
        
        # 3. Cost Breakdown Sheet
        self._create_cost_breakdown_sheet(wb, bom, assessment)
        
        # 4. BOM Lines Sheet
        self._create_bom_lines_sheet(wb, bom, assessment)
        
        # Save workbook
        wb.save(output_file)
    
    def _create_summary_sheet(self, wb, bom, assessment):
        """Create executive summary sheet"""
        ws = wb.create_sheet("Executive Summary", 0)
        
        # Header styling
        header_font = Font(size=16, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Title
        ws.merge_cells("A1:F1")
        ws["A1"] = "VM Assessment - Bill of Materials Report"
        ws["A1"].font = header_font
        ws["A1"].fill = header_fill
        ws["A1"].alignment = Alignment(horizontal="center")
        
        # Summary metrics
        ws["A3"] = "Report Generated:"
        ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        ws["A4"] = "Total VMs:"
        ws["B4"] = len(assessment.vms)
        
        ws["A5"] = "Powered On VMs:"
        ws["B5"] = len([vm for vm in assessment.vms if vm.is_powered_on])
        
        ws["A6"] = "Total Monthly Cost:"
        ws["B6"] = f"€{bom.total_monthly_cost:.2f}"
        
        ws["A7"] = "Currency:"
        ws["B7"] = bom.currency
        
        # OS Distribution - merge Unix into Linux
        ws["A9"] = "Operating System Distribution:"
        os_counts = {}
        for vm in assessment.vms:
            os_name = vm.os_type.value if vm.os_type else "Unknown"
            # Merge Unix systems into Linux category
            if os_name == "Unix":
                os_name = "Linux"
            os_counts[os_name] = os_counts.get(os_name, 0) + 1
        
        row = 10
        for os_name, count in os_counts.items():
            ws[f"A{row}"] = f"  {os_name}:"
            ws[f"B{row}"] = count
            row += 1
        
        # Cost breakdown by component
        ws[f"A{row + 1}"] = "Cost Breakdown by Component:"
        component_costs = {}
        for line in bom.line_items:
            comp_type = line.component_type
            component_costs[comp_type] = component_costs.get(comp_type, 0) + line.total_cost
        
        row += 2
        for comp_type, cost in component_costs.items():
            ws[f"A{row}"] = f"  {comp_type}:"
            ws[f"B{row}"] = f"€{cost:.2f}"
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_vm_details_sheet(self, wb, bom, assessment):
        """Create detailed VM information sheet"""
        ws = wb.create_sheet("VM Details")
        
        # Headers
        headers = [
            "VM Name", "OS Type", "Power State", "CPU Cores", "Memory (GB)", 
            "Storage (GB)", "Network Adapters", "Monthly Cost (€)"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        
        # VM data
        for row, vm in enumerate(assessment.vms, 2):
            ws.cell(row=row, column=1, value=vm.vm_name)
            ws.cell(row=row, column=2, value=vm.os_type.value if vm.os_type else "Unknown")
            ws.cell(row=row, column=3, value="Powered On" if vm.is_powered_on else "Powered Off")
            ws.cell(row=row, column=4, value=vm.cpu_cores)
            ws.cell(row=row, column=5, value=vm.memory_gb)
            ws.cell(row=row, column=6, value=vm.total_storage_gb)
            ws.cell(row=row, column=7, value=len(vm.networks))
            
            # Calculate VM monthly cost
            vm_cost = sum(line.total_cost for line in bom.line_items if line.vm_name == vm.vm_name)
            ws.cell(row=row, column=8, value=vm_cost)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_cost_breakdown_sheet(self, wb, bom, assessment):
        """Create cost breakdown sheet"""
        ws = wb.create_sheet("Cost Breakdown")
        
        # Headers
        headers = [
            "VM Name", "Component Type", "Description", "Quantity", 
            "Unit", "Unit Price (€)", "Total Cost (€)"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        
        # Cost data
        for row, line in enumerate(bom.line_items, 2):
            ws.cell(row=row, column=1, value=line.vm_name)
            ws.cell(row=row, column=2, value=line.component_type)
            ws.cell(row=row, column=3, value=line.description)
            ws.cell(row=row, column=4, value=line.quantity)
            ws.cell(row=row, column=5, value=line.unit)
            ws.cell(row=row, column=6, value=line.unit_price)
            ws.cell(row=row, column=7, value=line.total_cost)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_bom_lines_sheet(self, wb, bom, assessment):
        """Create detailed BOM lines sheet"""
        ws = wb.create_sheet("BOM Lines")
        
        # Headers
        headers = [
            "VM Name", "Component", "Description", "Qty", "Unit", 
            "Unit Price", "Total", "Pricing Model", "Notes"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D4E6F1", end_color="D4E6F1", fill_type="solid")
        
        # BOM data grouped by VM
        current_vm = None
        row = 2
        
        # Sort lines by VM name for better organization
        sorted_lines = sorted(bom.line_items, key=lambda x: x.vm_name)
        
        for line in sorted_lines:
            if current_vm != line.vm_name:
                if current_vm is not None:
                    # Add VM subtotal
                    ws.merge_cells(f"A{row}:F{row}")
                    vm_total = sum(l.total_cost for l in bom.line_items if l.vm_name == current_vm)
                    ws[f"A{row}"] = f"VM SUBTOTAL:"
                    ws[f"G{row}"] = f"€{vm_total:.2f}"
                    ws[f"A{row}"].font = Font(bold=True)
                    ws[f"G{row}"].font = Font(bold=True)
                    row += 1
                    
                    # Add separator
                    for col in range(1, 10):
                        ws.cell(row=row, column=col, value="-" * 20)
                    row += 1
                
                current_vm = line.vm_name
            
            ws.cell(row=row, column=1, value=line.vm_name)
            ws.cell(row=row, column=2, value=line.component_type)
            ws.cell(row=row, column=3, value=line.description)
            ws.cell(row=row, column=4, value=f"{line.quantity:.2f}")
            ws.cell(row=row, column=5, value=line.unit)
            ws.cell(row=row, column=6, value=f"€{line.unit_price:.4f}")
            ws.cell(row=row, column=7, value=f"€{line.total_cost:.2f}")
            ws.cell(row=row, column=8, value=getattr(line, 'pricing_model', 'on-demand'))
            ws.cell(row=row, column=9, value=getattr(line, 'notes', ''))
            row += 1
        
        # Final VM subtotal
        if current_vm:
            ws.merge_cells(f"A{row}:F{row}")
            vm_total = sum(l.total_cost for l in bom.line_items if l.vm_name == current_vm)
            ws[f"A{row}"] = f"VM SUBTOTAL:"
            ws[f"G{row}"] = f"€{vm_total:.2f}"
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"G{row}"].font = Font(bold=True)
            row += 2
        
        # Grand total
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = "GRAND TOTAL:"
        ws[f"G{row}"] = f"€{bom.total_monthly_cost:.2f}"
        ws[f"A{row}"].font = Font(bold=True, size=14)
        ws[f"G{row}"].font = Font(bold=True, size=14)
        ws[f"A{row}"].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        ws[f"G{row}"].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _generate_text_report(self, bom: BillOfMaterials, assessment: VMAssessment, output_file: Path):
        """Generate professional-quality text report using rich and tabulate libraries"""
        if FORMATTING_AVAILABLE:
            self._generate_rich_text_report(bom, assessment, output_file)
        else:
            self._generate_fallback_text_report(bom, assessment, output_file)
    
    def _generate_rich_text_report(self, bom: BillOfMaterials, assessment: VMAssessment, output_file: Path):
        """Generate high-quality text report using Rich library"""
        console = Console(file=StringIO(), width=120, force_terminal=False)
        
        # Header Panel
        header_text = Text("VM ASSESSMENT - BILL OF MATERIALS REPORT", style="bold white")
        metadata_text = Text(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Currency: {bom.currency} | Source: RVTools Export", style="dim")
        header_panel = Panel(
            Align.center(Text.assemble(header_text, "\n", metadata_text)),
            style="blue",
            box=box.DOUBLE
        )
        console.print(header_panel)
        console.print()
        
        # Executive Summary
        powered_on = len([vm for vm in assessment.vms if vm.is_powered_on])
        powered_off = len(assessment.vms) - powered_on
        
        summary_table = Table(title="Executive Summary", box=box.ROUNDED, show_header=False)
        summary_table.add_column("Metric", style="cyan")
        summary_table.add_column("Value", style="bold green")
        
        summary_table.add_row("📊 Total Virtual Machines", f"{len(assessment.vms):,}")
        summary_table.add_row("   ├─ Powered ON", f"{powered_on:,} VMs")
        summary_table.add_row("   └─ Powered OFF", f"{powered_off:,} VMs")
        summary_table.add_row("💰 Total Monthly Cost", f"€{bom.total_monthly_cost:,.2f}")
        summary_table.add_row("💰 Total Annual Cost", f"€{bom.total_monthly_cost * 12:,.2f}")
        
        console.print(summary_table)
        console.print()
        
        # OS Distribution - merge Unix into Linux
        os_counts = {}
        for vm in assessment.vms:
            os_name = vm.os_type.value if vm.os_type else "Unknown"
            # Merge Unix systems into Linux category
            if os_name == "Unix":
                os_name = "Linux"
            os_counts[os_name] = os_counts.get(os_name, 0) + 1
        
        os_table = Table(title="🖥️ Operating System Distribution", box=box.SIMPLE, show_footer=True)
        os_table.add_column("OS Type", style="cyan", footer="[bold]TOTAL")
        os_table.add_column("Count", justify="right", style="magenta", footer=f"[bold]{len(assessment.vms):,}")
        os_table.add_column("Percentage", justify="right", style="green", footer="[bold]100.0%")
        
        # Sort by count (descending) for better organization
        sorted_os = sorted(os_counts.items(), key=lambda x: x[1], reverse=True)
        
        for os_name, count in sorted_os:
            percentage = (count / len(assessment.vms)) * 100 if assessment.vms else 0
            os_table.add_row(os_name, f"{count:,}", f"{percentage:.1f}%")
        
        console.print(os_table)
        console.print()
        
        # Component Cost Breakdown
        component_costs = {}
        for line in bom.line_items:
            comp_type = line.component_type
            component_costs[comp_type] = component_costs.get(comp_type, 0) + line.total_cost
        
        component_table = Table(title="💼 Cost Breakdown by Component", box=box.SIMPLE)
        component_table.add_column("Component Type", style="cyan")
        component_table.add_column("Monthly Cost", justify="right", style="green")
        component_table.add_column("% of Total", justify="right", style="yellow")
        
        for comp_type, cost in sorted(component_costs.items(), key=lambda x: x[1], reverse=True):
            percentage = (cost / bom.total_monthly_cost) * 100 if bom.total_monthly_cost else 0
            component_table.add_row(comp_type, f"€{cost:,.2f}", f"{percentage:.1f}%")
        
        console.print(component_table)
        console.print()
        
        # Detailed VM Breakdown
        vm_lines = {}
        vm_totals = {}
        for line in bom.line_items:
            if line.vm_name not in vm_lines:
                vm_lines[line.vm_name] = []
                vm_totals[line.vm_name] = 0
            vm_lines[line.vm_name].append(line)
            vm_totals[line.vm_name] += line.total_cost
        
        # Sort VMs by cost (descending)
        sorted_vms = sorted(vm_totals.items(), key=lambda x: x[1], reverse=True)
        
        console.print(Text("💻 DETAILED VM COST BREAKDOWN", style="bold blue"))
        console.print()
        
        for i, (vm_name, vm_total) in enumerate(sorted_vms):
            lines = vm_lines[vm_name]
            vm = next((v for v in assessment.vms if v.vm_name == vm_name), None)
            
            # VM Header
            os_type = vm.os_type.value if vm and vm.os_type else "Unknown"
            power_state = "🟢 POWERED ON" if vm and vm.is_powered_on else "🔴 POWERED OFF"
            cpu_mem = f"{vm.cpu_cores} vCPU / {vm.memory_gb:.0f} GB RAM" if vm else "N/A"
            
            vm_header = f"#{i+1:>2} {vm_name} | {power_state} | {os_type} | {cpu_mem} | Monthly: €{vm_total:,.2f}"
            console.print(Text(vm_header, style="bold white on blue"))
            
            # VM Component Table
            vm_table = Table(box=box.SIMPLE_HEAD)
            vm_table.add_column("Component Type", style="cyan")
            vm_table.add_column("Description", style="white")
            vm_table.add_column("Quantity", justify="right", style="magenta")
            vm_table.add_column("Unit", style="yellow")
            vm_table.add_column("Cost (€)", justify="right", style="green")
            
            for line in lines:
                vm_table.add_row(
                    line.component_type,
                    line.description,
                    f"{line.quantity:.1f}",
                    line.unit,
                    f"{line.total_cost:.2f}"
                )
            
            # Add subtotal row
            vm_table.add_row("", "", "", "[bold]SUBTOTAL", f"[bold green]€{vm_total:,.2f}")
            
            console.print(vm_table)
            console.print()
        
        # Grand Total
        total_panel = Panel(
            Align.center(Text(f"🎯 TOTAL MONTHLY COST: €{bom.total_monthly_cost:,.2f}\n🎯 TOTAL ANNUAL COST: €{bom.total_monthly_cost * 12:,.2f}", style="bold white")),
            style="green",
            box=box.DOUBLE
        )
        console.print(total_panel)
        console.print()
        
        # Footer
        console.print(Text("Generated by RVTools BOM Assessment Tool", style="dim italic"))
        console.print(Text(f"Oracle Cloud Infrastructure Pricing - {datetime.now().strftime('%Y-%m-%d')}", style="dim italic"))
        
        # Write to file
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(console.file.getvalue())
    
    def _generate_fallback_text_report(self, bom: BillOfMaterials, assessment: VMAssessment, output_file: Path):
        """Fallback text report using tabulate library if Rich is not available"""
        with open(output_file, 'w', encoding='utf-8') as f:
            # Header
            f.write("=" * 100 + "\n")
            f.write("VM ASSESSMENT - BILL OF MATERIALS REPORT".center(100) + "\n")
            f.write("=" * 100 + "\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Currency: {bom.currency} | Source: RVTools Export\n\n")
            
            # Executive Summary using tabulate
            powered_on = len([vm for vm in assessment.vms if vm.is_powered_on])
            powered_off = len(assessment.vms) - powered_on
            
            summary_data = [
                ["Total Virtual Machines", f"{len(assessment.vms):,}"],
                ["├─ Powered ON", f"{powered_on:,} VMs"],
                ["└─ Powered OFF", f"{powered_off:,} VMs"],
                ["Total Monthly Cost", f"€{bom.total_monthly_cost:,.2f}"],
                ["Total Annual Cost", f"€{bom.total_monthly_cost * 12:,.2f}"]
            ]
            
            f.write("EXECUTIVE SUMMARY\n")
            f.write(tabulate(summary_data, tablefmt="grid", colalign=("left", "right")))
            f.write("\n\n")
            
            # OS Distribution - merge Unix into Linux
            os_counts = {}
            for vm in assessment.vms:
                os_name = vm.os_type.value if vm.os_type else "Unknown"
                # Merge Unix systems into Linux category
                if os_name == "Unix":
                    os_name = "Linux"
                os_counts[os_name] = os_counts.get(os_name, 0) + 1
            
            os_data = []
            # Sort by count (descending) for better organization
            sorted_os = sorted(os_counts.items(), key=lambda x: x[1], reverse=True)
            
            for os_name, count in sorted_os:
                percentage = (count / len(assessment.vms)) * 100 if assessment.vms else 0
                os_data.append([os_name, f"{count:,}", f"{percentage:.1f}%"])
            
            # Add totals row
            os_data.append(["", "", ""])
            os_data.append(["TOTAL", f"{len(assessment.vms):,}", "100.0%"])
            
            f.write("OPERATING SYSTEM DISTRIBUTION\n")
            f.write(tabulate(os_data, headers=["OS Type", "Count", "Percentage"], tablefmt="grid"))
            f.write("\n\n")
            
            # Component Cost Breakdown
            component_costs = {}
            for line in bom.line_items:
                comp_type = line.component_type
                component_costs[comp_type] = component_costs.get(comp_type, 0) + line.total_cost
            
            component_data = []
            for comp_type, cost in sorted(component_costs.items(), key=lambda x: x[1], reverse=True):
                percentage = (cost / bom.total_monthly_cost) * 100 if bom.total_monthly_cost else 0
                component_data.append([comp_type, f"€{cost:,.2f}", f"{percentage:.1f}%"])
            
            f.write("COST BREAKDOWN BY COMPONENT\n")
            f.write(tabulate(component_data, headers=["Component Type", "Monthly Cost", "% of Total"], tablefmt="grid"))
            f.write("\n\n")
            
            # Detailed VM Breakdown
            vm_lines = {}
            vm_totals = {}
            for line in bom.line_items:
                if line.vm_name not in vm_lines:
                    vm_lines[line.vm_name] = []
                    vm_totals[line.vm_name] = 0
                vm_lines[line.vm_name].append(line)
                vm_totals[line.vm_name] += line.total_cost
            
            # Sort VMs by cost (descending)
            sorted_vms = sorted(vm_totals.items(), key=lambda x: x[1], reverse=True)
            
            f.write("DETAILED VM COST BREAKDOWN\n")
            f.write("=" * 100 + "\n")
            
            for i, (vm_name, vm_total) in enumerate(sorted_vms):
                lines = vm_lines[vm_name]
                vm = next((v for v in assessment.vms if v.vm_name == vm_name), None)
                
                # VM Header
                os_type = vm.os_type.value if vm and vm.os_type else "Unknown"
                power_state = "POWERED ON" if vm and vm.is_powered_on else "POWERED OFF"
                cpu_mem = f"{vm.cpu_cores} vCPU / {vm.memory_gb:.0f} GB RAM" if vm else "N/A"
                
                f.write(f"\n#{i+1:>2} {vm_name} | {power_state} | {os_type} | {cpu_mem} | Monthly: €{vm_total:,.2f}\n")
                f.write("-" * 100 + "\n")
                
                # VM Component Table
                vm_data = []
                for line in lines:
                    vm_data.append([
                        line.component_type,
                        line.description,
                        f"{line.quantity:.1f}",
                        line.unit,
                        f"€{line.total_cost:.2f}"
                    ])
                
                # Add subtotal
                vm_data.append(["", "", "", "SUBTOTAL", f"€{vm_total:,.2f}"])
                
                f.write(tabulate(vm_data, headers=["Component", "Description", "Qty", "Unit", "Cost"], tablefmt="grid"))
                f.write("\n")
            
            # Grand Total
            f.write("\n" + "=" * 100 + "\n")
            f.write(f"TOTAL MONTHLY COST: €{bom.total_monthly_cost:,.2f}\n")
            f.write(f"TOTAL ANNUAL COST: €{bom.total_monthly_cost * 12:,.2f}\n")
            f.write("=" * 100 + "\n\n")
            
            # Footer
            f.write("Generated by RVTools BOM Assessment Tool\n")
            f.write(f"Oracle Cloud Infrastructure Pricing - {datetime.now().strftime('%Y-%m-%d')}\n")
    
    def _generate_csv_report(self, bom: BillOfMaterials, assessment: VMAssessment, output_file: Path):
        """Generate CSV report for analysis"""
        with open(output_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Headers
            writer.writerow([
                'VM_Name', 'OS_Type', 'Power_State', 'CPU_Cores', 'Memory_GB', 
                'Storage_GB', 'Network_Adapters', 'Component_Type', 'Component_Description',
                'Quantity', 'Unit', 'Unit_Price_EUR', 'Total_Cost_EUR', 'Pricing_Model'
            ])
            
            # Data rows
            for line in bom.line_items:
                # Find corresponding VM
                vm = next((v for v in assessment.vms if v.vm_name == line.vm_name), None)
                
                if vm:
                    writer.writerow([
                        line.vm_name,
                        vm.os_type.value if vm.os_type else 'Unknown',
                        'Powered_On' if vm.is_powered_on else 'Powered_Off',
                        vm.cpu_cores,
                        vm.memory_gb,
                        vm.total_storage_gb,
                        len(vm.networks),
                        line.component_type,
                        line.description,
                        line.quantity,
                        line.unit,
                        line.unit_price,
                        line.total_cost,
                        getattr(line, 'pricing_model', 'on-demand')
                    ])