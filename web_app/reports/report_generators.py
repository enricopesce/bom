"""
Report Generators for VM Assessments and Bills of Materials

Concrete implementations of ReportGenerator for different output formats
"""

import csv
import json
from typing import List, Dict, Any
from datetime import datetime
from pathlib import Path

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from processors.base_processor import ReportGenerator
from models.vm_models import VMAssessment, BillOfMaterials, OSType

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class CSVReportGenerator(ReportGenerator):
    """CSV report generator"""
    
    @property
    def format_name(self) -> str:
        return "CSV"
    
    @property
    def file_extension(self) -> str:
        return "csv"
    
    def generate_assessment_report(self, assessment: VMAssessment, output_file: str) -> str:
        """Generate CSV assessment report"""
        with open(output_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Header
            writer.writerow([f'VM Assessment Report - {assessment.source_format}'])
            writer.writerow([f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
            writer.writerow([f'Total VMs: {assessment.total_vms}'])
            writer.writerow([f'Powered On VMs: {assessment.powered_on_vms}'])
            writer.writerow([])
            
            # VM details header
            writer.writerow([
                'VM ID', 'VM Name', 'OS Type', 'OS Config', 'CPU Cores', 'Memory GB', 
                'Storage GB', 'Power State', 'Cluster', 'Host', 'Datacenter', 'Annotation'
            ])
            
            # VM data
            for vm in assessment.vms:
                writer.writerow([
                    vm.vm_id,
                    vm.vm_name,
                    vm.os_type.value,
                    vm.os_config or '',
                    vm.cpu_cores,
                    vm.memory_gb,
                    vm.total_storage_gb,
                    vm.power_state.value,
                    vm.cluster or '',
                    vm.host or '',
                    vm.datacenter or '',
                    vm.annotation or ''
                ])
        
        self.logger.info(f"Generated CSV assessment report: {output_file}")
        return output_file
    
    def generate_bom_report(self, bom: BillOfMaterials, output_file: str) -> str:
        """Generate CSV BOM report"""
        with open(output_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Header
            writer.writerow([f'Bill of Materials - {bom.pricing_source}'])
            writer.writerow([f'Currency: {bom.currency}'])
            writer.writerow([f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
            writer.writerow([])
            
            # BOM header
            writer.writerow([
                'VM ID', 'VM Name', 'OS Type', 'Component Type', 'Description',
                'Quantity', 'Unit', f'Unit Price ({bom.currency})', f'Total Cost ({bom.currency})'
            ])
            
            # BOM line items
            for item in bom.line_items:
                writer.writerow([
                    item.vm_id,
                    item.vm_name,
                    item.os_type.value,
                    item.component_type,
                    item.description,
                    item.quantity,
                    item.unit,
                    item.unit_price,
                    item.total_cost
                ])
            
            # Summary
            writer.writerow([])
            writer.writerow(['SUMMARY'])
            writer.writerow([f'Total Monthly Cost ({bom.currency})', '', '', '', '', '', '', '', bom.total_monthly_cost])
            writer.writerow([f'Total Annual Cost ({bom.currency})', '', '', '', '', '', '', '', bom.total_annual_cost])
        
        self.logger.info(f"Generated CSV BOM report: {output_file}")
        return output_file


class ExcelReportGenerator(ReportGenerator):
    """Excel report generator"""
    
    @property
    def format_name(self) -> str:
        return "Excel"
    
    @property
    def file_extension(self) -> str:
        return "xlsx"
    
    def generate_assessment_report(self, assessment: VMAssessment, output_file: str) -> str:
        """Generate Excel assessment report"""
        if not EXCEL_AVAILABLE:
            self.logger.warning("openpyxl not available, skipping Excel report")
            return ""
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "VM Assessment"
        
        # Header with formatting
        header_font = Font(bold=True, size=14)
        ws['A1'] = f'VM Assessment Report - {assessment.source_format}'
        ws['A1'].font = header_font
        
        ws['A3'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws['A4'] = f'Total VMs: {assessment.total_vms}'
        ws['A5'] = f'Powered On VMs: {assessment.powered_on_vms}'
        
        # Column headers
        headers = ['VM ID', 'VM Name', 'OS Type', 'CPU Cores', 'Memory GB', 'Storage GB', 'Power State', 'Cluster']
        header_row = 7
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True)
        
        # VM data
        for row, vm in enumerate(assessment.vms, header_row + 1):
            ws.cell(row=row, column=1, value=vm.vm_id)
            ws.cell(row=row, column=2, value=vm.vm_name)
            ws.cell(row=row, column=3, value=vm.os_type.value)
            ws.cell(row=row, column=4, value=vm.cpu_cores)
            ws.cell(row=row, column=5, value=vm.memory_gb)
            ws.cell(row=row, column=6, value=vm.total_storage_gb)
            ws.cell(row=row, column=7, value=vm.power_state.value)
            ws.cell(row=row, column=8, value=vm.cluster or '')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 12), 50)
        
        wb.save(output_file)
        self.logger.info(f"Generated Excel assessment report: {output_file}")
        return output_file
    
    def generate_bom_report(self, bom: BillOfMaterials, output_file: str) -> str:
        """Generate Excel BOM report"""
        if not EXCEL_AVAILABLE:
            self.logger.warning("openpyxl not available, skipping Excel report")
            return ""
        
        wb = openpyxl.Workbook()
        
        # Summary sheet
        summary_ws = wb.active
        summary_ws.title = "Summary"
        
        header_font = Font(bold=True, size=14)
        summary_ws['A1'] = f'Bill of Materials Summary'
        summary_ws['A1'].font = header_font
        
        summary_ws['A3'] = f'Pricing Source: {bom.pricing_source}'
        summary_ws['A4'] = f'Currency: {bom.currency}'
        summary_ws['A5'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        
        summary_ws['A7'] = f'Total Monthly Cost: {bom.currency} {bom.total_monthly_cost:,.2f}'
        summary_ws['A8'] = f'Total Annual Cost: {bom.currency} {bom.total_annual_cost:,.2f}'
        
        # Cost breakdown by component
        summary_ws['A10'] = 'Cost Breakdown by Component'
        summary_ws['A10'].font = Font(bold=True)
        
        component_costs = bom.get_cost_by_component()
        row = 11
        for component, cost in sorted(component_costs.items(), key=lambda x: x[1], reverse=True):
            percentage = (cost / bom.total_monthly_cost) * 100
            summary_ws[f'A{row}'] = component
            summary_ws[f'B{row}'] = f'{bom.currency} {cost:,.2f}'
            summary_ws[f'C{row}'] = f'{percentage:.1f}%'
            row += 1
        
        # Detailed BOM sheet
        detail_ws = wb.create_sheet("Detailed BOM")
        
        headers = ['VM Name', 'OS Type', 'Component', 'Description', 'Quantity', 'Unit', f'Unit Price ({bom.currency})', f'Total Cost ({bom.currency})']
        
        for col, header in enumerate(headers, 1):
            cell = detail_ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        # BOM line items
        for row, item in enumerate(bom.line_items, 2):
            detail_ws.cell(row=row, column=1, value=item.vm_name)
            detail_ws.cell(row=row, column=2, value=item.os_type.value)
            detail_ws.cell(row=row, column=3, value=item.component_type)
            detail_ws.cell(row=row, column=4, value=item.description)
            detail_ws.cell(row=row, column=5, value=item.quantity)
            detail_ws.cell(row=row, column=6, value=item.unit)
            detail_ws.cell(row=row, column=7, value=item.unit_price)
            detail_ws.cell(row=row, column=8, value=item.total_cost)
        
        # Auto-adjust column widths for both sheets
        for ws in [summary_ws, detail_ws]:
            for column in ws.columns:
                max_length = max(len(str(cell.value or '')) for cell in column)
                ws.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 12), 50)
        
        wb.save(output_file)
        self.logger.info(f"Generated Excel BOM report: {output_file}")
        return output_file


class SalesExcelReportGenerator(ReportGenerator):
    """Sales-focused Excel report generator with business-friendly formatting"""
    
    @property
    def format_name(self) -> str:
        return "Sales Excel"
    
    @property
    def file_extension(self) -> str:
        return "xlsx"
    
    def generate_assessment_report(self, assessment: VMAssessment, output_file: str) -> str:
        """Generate sales-focused Excel assessment report"""
        if not EXCEL_AVAILABLE:
            self.logger.warning("openpyxl not available, skipping Sales Excel report")
            return ""
        
        wb = openpyxl.Workbook()
        
        # Remove default sheet and create custom sheets
        wb.remove(wb.active)
        
        # 1. Executive Summary Sheet
        exec_ws = wb.create_sheet("Executive Summary")
        self._create_executive_summary_sheet(exec_ws, assessment)
        
        # 2. Infrastructure Overview Sheet  
        infra_ws = wb.create_sheet("Infrastructure Overview")
        self._create_infrastructure_overview_sheet(infra_ws, assessment)
        
        # 3. VM Inventory Sheet
        vm_ws = wb.create_sheet("VM Inventory")
        self._create_vm_inventory_sheet(vm_ws, assessment)
        
        wb.save(output_file)
        self.logger.info(f"Generated Sales Excel assessment report: {output_file}")
        return output_file
    
    def generate_bom_report(self, bom: BillOfMaterials, output_file: str) -> str:
        """Generate sales-focused Excel BOM report"""
        if not EXCEL_AVAILABLE:
            self.logger.warning("openpyxl not available, skipping Sales Excel report")
            return ""
        
        wb = openpyxl.Workbook()
        
        # Remove default sheet and create custom sheets
        wb.remove(wb.active)
        
        # 1. Cost Summary Sheet
        cost_ws = wb.create_sheet("Cost Summary")
        self._create_cost_summary_sheet(cost_ws, bom)
        
        # 2. Monthly Cost Analysis Sheet
        monthly_ws = wb.create_sheet("Monthly Breakdown")
        self._create_monthly_breakdown_sheet(monthly_ws, bom)
        
        # 3. Savings Opportunities Sheet
        savings_ws = wb.create_sheet("Savings Opportunities")
        self._create_savings_opportunities_sheet(savings_ws, bom)
        
        # 4. Detailed Pricing Sheet
        detail_ws = wb.create_sheet("Detailed Pricing")
        self._create_detailed_pricing_sheet(detail_ws, bom)
        
        wb.save(output_file)
        self.logger.info(f"Generated Sales Excel BOM report: {output_file}")
        return output_file
    
    def _create_executive_summary_sheet(self, ws, assessment):
        """Create executive summary sheet for sales presentation"""
        # Title and header
        ws['A1'] = 'VM Assessment - Executive Summary'
        ws['A1'].font = Font(bold=True, size=16, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
        ws.merge_cells('A1:F1')
        
        # Key metrics section
        ws['A3'] = 'Key Infrastructure Metrics'
        ws['A3'].font = Font(bold=True, size=14)
        
        stats = assessment.get_summary_stats()
        
        # Metrics table
        metrics = [
            ['Total Virtual Machines', stats['total_vms']],
            ['Active VMs (Powered On)', stats['powered_on_vms']],
            ['Inactive VMs (Powered Off)', stats['powered_off_vms']],
            ['Total CPU Cores', stats['total_cpu_cores']],
            ['Total Memory (GB)', f"{stats['total_memory_gb']:.0f}"],
            ['Total Storage (TB)', f"{stats['total_storage_gb']/1000:.1f}"],
        ]
        
        for i, (metric, value) in enumerate(metrics, 5):
            ws[f'A{i}'] = metric
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
        
        # OS Distribution
        ws['D3'] = 'Operating System Distribution'
        ws['D3'].font = Font(bold=True, size=14)
        
        os_dist = stats['os_distribution']
        for i, (os_type, count) in enumerate(os_dist.items(), 5):
            percentage = (count / stats['total_vms']) * 100
            ws[f'D{i}'] = os_type
            ws[f'E{i}'] = count
            ws[f'F{i}'] = f"{percentage:.1f}%"
            ws[f'D{i}'].font = Font(bold=True)
        
        # Styling
        for row in ws['A3:F10']:
            for cell in row:
                cell.alignment = Alignment(vertical='center')
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _create_infrastructure_overview_sheet(self, ws, assessment):
        """Create infrastructure overview for technical discussion"""
        ws['A1'] = 'Infrastructure Overview'
        ws['A1'].font = Font(bold=True, size=16, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
        ws.merge_cells('A1:H1')
        
        # Headers
        headers = ['VM Name', 'OS Type', 'CPU Cores', 'Memory (GB)', 'Storage (GB)', 'Power State', 'Cluster', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # VM data
        row = 4
        for vm in sorted(assessment.vms, key=lambda x: x.vm_name):
            ws.cell(row=row, column=1, value=vm.vm_name)
            ws.cell(row=row, column=2, value=vm.os_type.value)
            ws.cell(row=row, column=3, value=vm.cpu_cores)
            ws.cell(row=row, column=4, value=vm.memory_gb)
            ws.cell(row=row, column=5, value=vm.total_storage_gb)
            ws.cell(row=row, column=6, value=vm.power_state.value)
            ws.cell(row=row, column=7, value=vm.cluster or 'N/A')
            ws.cell(row=row, column=8, value=vm.annotation or '')
            
            # Color code by power state
            if vm.power_state.value == 'poweredOn':
                fill_color = 'E2EFDA'  # Light green
            else:
                fill_color = 'FCE4D6'  # Light orange
            
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            
            row += 1
        
        self._auto_adjust_columns(ws)
    
    def _create_vm_inventory_sheet(self, ws, assessment):
        """Create detailed VM inventory sheet"""
        ws['A1'] = 'Virtual Machine Inventory'
        ws['A1'].font = Font(bold=True, size=16)
        
        # Summary stats
        stats = assessment.get_summary_stats()
        ws['A3'] = f"Total VMs: {stats['total_vms']}"
        ws['B3'] = f"Powered On: {stats['powered_on_vms']}"
        ws['C3'] = f"Powered Off: {stats['powered_off_vms']}"
        
        # Powered On VMs
        ws['A5'] = 'Active Virtual Machines (Powered On)'
        ws['A5'].font = Font(bold=True, size=14, color='FFFFFF')
        ws['A5'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        ws.merge_cells('A5:F5')
        
        headers = ['VM Name', 'OS Type', 'CPU Cores', 'Memory (GB)', 'Storage (GB)', 'Cluster']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = Font(bold=True)
        
        row = 7
        powered_on_vms = [vm for vm in assessment.vms if vm.power_state.value == 'poweredOn']
        for vm in sorted(powered_on_vms, key=lambda x: x.vm_name):
            ws.cell(row=row, column=1, value=vm.vm_name)
            ws.cell(row=row, column=2, value=vm.os_type.value)
            ws.cell(row=row, column=3, value=vm.cpu_cores)
            ws.cell(row=row, column=4, value=vm.memory_gb)
            ws.cell(row=row, column=5, value=vm.total_storage_gb)
            ws.cell(row=row, column=6, value=vm.cluster or 'N/A')
            row += 1
        
        # Powered Off VMs
        powered_off_vms = [vm for vm in assessment.vms if vm.power_state.value != 'poweredOn']
        if powered_off_vms:
            row += 2
            ws[f'A{row}'] = 'Inactive Virtual Machines (Powered Off)'
            ws[f'A{row}'].font = Font(bold=True, size=14, color='FFFFFF')
            ws[f'A{row}'].fill = PatternFill(start_color='C55A5A', end_color='C55A5A', fill_type='solid')
            ws.merge_cells(f'A{row}:F{row}')
            
            row += 1
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
            
            row += 1
            for vm in sorted(powered_off_vms, key=lambda x: x.vm_name):
                ws.cell(row=row, column=1, value=vm.vm_name)
                ws.cell(row=row, column=2, value=vm.os_type.value)
                ws.cell(row=row, column=3, value=vm.cpu_cores)
                ws.cell(row=row, column=4, value=vm.memory_gb)
                ws.cell(row=row, column=5, value=vm.total_storage_gb)
                ws.cell(row=row, column=6, value=vm.cluster or 'N/A')
                row += 1
        
        self._auto_adjust_columns(ws)
    
    def _create_cost_summary_sheet(self, ws, bom):
        """Create executive cost summary for sales presentation"""
        # Title
        ws['A1'] = f'Cost Summary - {bom.pricing_source}'
        ws['A1'].font = Font(bold=True, size=16, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
        ws.merge_cells('A1:D1')
        
        # Key cost metrics
        ws['A3'] = 'Investment Summary'
        ws['A3'].font = Font(bold=True, size=14)
        
        cost_summary = [
            ['Monthly Cloud Cost', f"{bom.currency} {bom.total_monthly_cost:,.2f}"],
            ['Annual Cloud Cost', f"{bom.currency} {bom.total_annual_cost:,.2f}"],
            ['3-Year Total Cost', f"{bom.currency} {bom.total_annual_cost * 3:,.2f}"],
            ['Average Cost per VM/Month', f"{bom.currency} {bom.total_monthly_cost / len(set(item.vm_name for item in bom.line_items)):,.2f}"],
        ]
        
        for i, (metric, value) in enumerate(cost_summary, 5):
            ws[f'A{i}'] = metric
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
            if 'Annual' in metric or '3-Year' in metric:
                ws[f'B{i}'].font = Font(bold=True, color='C55A5A')
        
        # Cost breakdown pie chart data
        ws['A10'] = 'Cost Breakdown by Component'
        ws['A10'].font = Font(bold=True, size=14)
        
        component_costs = bom.get_cost_by_component()
        row = 12
        for component, cost in sorted(component_costs.items(), key=lambda x: x[1], reverse=True):
            percentage = (cost / bom.total_monthly_cost) * 100
            ws[f'A{row}'] = component
            ws[f'B{row}'] = f"{bom.currency} {cost:,.2f}"
            ws[f'C{row}'] = f"{percentage:.1f}%"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Headers for breakdown
        ws['A11'] = 'Component'
        ws['B11'] = 'Monthly Cost'
        ws['C11'] = 'Percentage'
        for col in ['A11', 'B11', 'C11']:
            ws[col].font = Font(bold=True)
        
        self._auto_adjust_columns(ws)
    
    def _create_monthly_breakdown_sheet(self, ws, bom):
        """Create monthly cost breakdown by VM"""
        ws['A1'] = 'Monthly Cost Breakdown by Virtual Machine'
        ws['A1'].font = Font(bold=True, size=16)
        
        # Headers
        headers = ['VM Name', 'OS Type', 'Monthly Cost', 'Annual Cost', 'Primary Cost Driver']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        # Calculate VM costs
        vm_costs = {}
        for item in bom.line_items:
            if item.vm_name not in vm_costs:
                vm_costs[item.vm_name] = {
                    'os_type': item.os_type.value,
                    'total_cost': 0,
                    'components': {}
                }
            vm_costs[item.vm_name]['total_cost'] += item.total_cost
            vm_costs[item.vm_name]['components'][item.component_type] = item.total_cost
        
        # Sort by cost (highest first)
        sorted_vms = sorted(vm_costs.items(), key=lambda x: x[1]['total_cost'], reverse=True)
        
        row = 4
        for vm_name, data in sorted_vms:
            # Find primary cost driver
            primary_driver = max(data['components'].items(), key=lambda x: x[1])[0]
            
            ws.cell(row=row, column=1, value=vm_name)
            ws.cell(row=row, column=2, value=data['os_type'])
            ws.cell(row=row, column=3, value=data['total_cost'])
            ws.cell(row=row, column=4, value=data['total_cost'] * 12)
            ws.cell(row=row, column=5, value=primary_driver)
            
            # Format currency cells
            ws.cell(row=row, column=3).number_format = f'"{bom.currency}" #,##0.00'
            ws.cell(row=row, column=4).number_format = f'"{bom.currency}" #,##0.00'
            
            # Color code by cost level
            if data['total_cost'] > bom.total_monthly_cost * 0.1:  # High cost (>10% of total)
                fill_color = 'FCE4D6'  # Light red
            elif data['total_cost'] > bom.total_monthly_cost * 0.05:  # Medium cost (>5% of total)
                fill_color = 'FFF2CC'  # Light yellow
            else:
                fill_color = 'E2EFDA'  # Light green
            
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            
            row += 1
        
        # Total row
        ws.cell(row=row, column=1, value='TOTAL')
        ws.cell(row=row, column=3, value=bom.total_monthly_cost)
        ws.cell(row=row, column=4, value=bom.total_annual_cost)
        for col in range(1, 6):
            ws.cell(row=row, column=col).font = Font(bold=True)
        
        self._auto_adjust_columns(ws)
    
    def _create_savings_opportunities_sheet(self, ws, bom):
        """Create savings opportunities analysis"""
        ws['A1'] = 'Potential Savings Opportunities'
        ws['A1'].font = Font(bold=True, size=16, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        ws.merge_cells('A1:D1')
        
        # Savings scenarios
        ws['A3'] = 'Cloud Migration Savings Analysis'
        ws['A3'].font = Font(bold=True, size=14)
        
        monthly_cost = bom.total_monthly_cost
        
        savings_scenarios = [
            ['Right-sizing VMs (10% reduction)', monthly_cost * 0.1, 'Optimize VM sizes based on actual usage'],
            ['Reserved Instances (20% discount)', monthly_cost * 0.2, 'Commit to 1-3 year terms for predictable workloads'],
            ['Automated Scaling (15% reduction)', monthly_cost * 0.15, 'Implement auto-scaling for variable workloads'],
            ['Storage Optimization (8% reduction)', monthly_cost * 0.08, 'Implement tiered storage and compression'],
            ['Eliminate Powered-Off VMs', 0, 'Remove or consolidate inactive virtual machines'],
        ]
        
        ws['A5'] = 'Optimization Opportunity'
        ws['B5'] = 'Monthly Savings'
        ws['C5'] = 'Annual Savings'
        ws['D5'] = 'Description'
        for col in ['A5', 'B5', 'C5', 'D5']:
            ws[col].font = Font(bold=True)
        
        total_savings = 0
        for i, (opportunity, monthly_saving, description) in enumerate(savings_scenarios, 6):
            ws[f'A{i}'] = opportunity
            ws[f'B{i}'] = f"{bom.currency} {monthly_saving:,.2f}"
            ws[f'C{i}'] = f"{bom.currency} {monthly_saving * 12:,.2f}"
            ws[f'D{i}'] = description
            total_savings += monthly_saving
        
        # Total savings
        row = len(savings_scenarios) + 7
        ws[f'A{row}'] = 'TOTAL POTENTIAL SAVINGS'
        ws[f'B{row}'] = f"{bom.currency} {total_savings:,.2f}"
        ws[f'C{row}'] = f"{bom.currency} {total_savings * 12:,.2f}"
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = Font(bold=True, color='FFFFFF')
            ws[f'{col}{row}'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        
        # ROI Analysis
        row += 3
        ws[f'A{row}'] = 'Return on Investment Analysis'
        ws[f'A{row}'].font = Font(bold=True, size=14)
        
        roi_data = [
            ['Original Annual Cost', f"{bom.currency} {bom.total_annual_cost:,.2f}"],
            ['Optimized Annual Cost', f"{bom.currency} {(bom.total_annual_cost - total_savings * 12):,.2f}"],
            ['Total Annual Savings', f"{bom.currency} {total_savings * 12:,.2f}"],
            ['ROI Percentage', f"{(total_savings * 12 / bom.total_annual_cost * 100):.1f}%"],
        ]
        
        for i, (metric, value) in enumerate(roi_data, row + 2):
            ws[f'A{i}'] = metric
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
        
        self._auto_adjust_columns(ws)
    
    def _create_detailed_pricing_sheet(self, ws, bom):
        """Create detailed pricing breakdown"""
        ws['A1'] = 'Detailed Pricing Breakdown'
        ws['A1'].font = Font(bold=True, size=16)
        
        # Headers
        headers = ['VM Name', 'OS Type', 'Component', 'Description', 'Quantity', 'Unit Price', 'Monthly Cost']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        # Group by VM for better presentation
        vm_items = {}
        for item in bom.line_items:
            if item.vm_name not in vm_items:
                vm_items[item.vm_name] = []
            vm_items[item.vm_name].append(item)
        
        # Sort VMs by total cost
        vm_totals = {vm_name: sum(item.total_cost for item in items) for vm_name, items in vm_items.items()}
        sorted_vms = sorted(vm_items.items(), key=lambda x: vm_totals[x[0]], reverse=True)
        
        row = 4
        for vm_name, items in sorted_vms:
            vm_total = sum(item.total_cost for item in items)
            
            for i, item in enumerate(items):
                # Show VM name only on first line
                vm_display = vm_name if i == 0 else ""
                os_display = item.os_type.value if i == 0 else ""
                
                ws.cell(row=row, column=1, value=vm_display)
                ws.cell(row=row, column=2, value=os_display)
                ws.cell(row=row, column=3, value=item.component_type)
                ws.cell(row=row, column=4, value=item.description)
                ws.cell(row=row, column=5, value=item.quantity)
                ws.cell(row=row, column=6, value=item.unit_price)
                ws.cell(row=row, column=7, value=item.total_cost)
                
                # Format currency cells
                ws.cell(row=row, column=6).number_format = f'"{bom.currency}" #,##0.0000'
                ws.cell(row=row, column=7).number_format = f'"{bom.currency}" #,##0.00'
                
                row += 1
            
            # VM subtotal
            ws.cell(row=row, column=1, value='VM SUBTOTAL')
            ws.cell(row=row, column=7, value=vm_total)
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=7).font = Font(bold=True)
            ws.cell(row=row, column=7).number_format = f'"{bom.currency}" #,##0.00'
            row += 1
        
        self._auto_adjust_columns(ws)
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths for better readability"""
        for column in ws.columns:
            max_length = 0
            try:
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if hasattr(cell, 'value') and cell.value is not None:
                            length = len(str(cell.value))
                            if length > max_length:
                                max_length = length
                    except:
                        pass
                adjusted_width = min(max(max_length + 2, 12), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            except:
                # Skip columns with merged cells or other issues
                pass


class TextReportGenerator(ReportGenerator):
    """Text report generator for console output and text files"""
    
    @property
    def format_name(self) -> str:
        return "Text"
    
    @property
    def file_extension(self) -> str:
        return "txt"
    
    def generate_assessment_report(self, assessment: VMAssessment, output_file: str) -> str:
        """Generate text assessment report"""
        lines = []
        lines.append("=" * 80)
        lines.append("VM ASSESSMENT REPORT")
        lines.append("=" * 80)
        lines.append(f"Source: {assessment.source_format}")
        lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"Assessment Date: {assessment.assessment_date}")
        lines.append("")
        
        # Summary
        lines.append("SUMMARY")
        lines.append("-" * 40)
        stats = assessment.get_summary_stats()
        lines.append(f"Total VMs: {stats['total_vms']}")
        lines.append(f"Powered On VMs: {stats['powered_on_vms']}")
        lines.append(f"Powered Off VMs: {stats['powered_off_vms']}")
        lines.append(f"Total CPU Cores: {stats['total_cpu_cores']}")
        lines.append(f"Total Memory: {stats['total_memory_gb']:.1f} GB")
        lines.append(f"Total Storage: {stats['total_storage_gb']:.1f} GB")
        lines.append("")
        
        # OS Distribution
        lines.append("OS DISTRIBUTION")
        lines.append("-" * 40)
        for os_type, count in stats['os_distribution'].items():
            percentage = (count / stats['total_vms']) * 100
            lines.append(f"{os_type}: {count} VMs ({percentage:.1f}%)")
        lines.append("")
        
        # VM Details
        lines.append("VM DETAILS")
        lines.append("-" * 80)
        lines.append(f"{'VM Name':<25} {'OS Type':<12} {'CPU':<4} {'RAM GB':<8} {'Disk GB':<10} {'Power':<10}")
        lines.append("-" * 80)
        
        for vm in assessment.vms:
            lines.append(f"{vm.vm_name:<25} {vm.os_type.value:<12} {vm.cpu_cores:<4} {vm.memory_gb:<8.1f} {vm.total_storage_gb:<10.1f} {vm.power_state.value:<10}")
        
        # Write to file
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        
        self.logger.info(f"Generated text assessment report: {output_file}")
        return output_file
    
    def generate_bom_report(self, bom: BillOfMaterials, output_file: str) -> str:
        """Generate text BOM report"""
        lines = []
        lines.append("=" * 120)
        lines.append("BILL OF MATERIALS REPORT")
        lines.append("=" * 120)
        lines.append(f"Pricing Source: {bom.pricing_source}")
        lines.append(f"Currency: {bom.currency}")
        lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("")
        
        # Cost Summary
        lines.append("COST SUMMARY")
        lines.append("-" * 60)
        lines.append(f"Total Monthly Cost: {bom.currency} {bom.total_monthly_cost:,.2f}")
        lines.append(f"Total Annual Cost: {bom.currency} {bom.total_annual_cost:,.2f}")
        lines.append("")
        
        # Component Breakdown
        lines.append("COST BREAKDOWN BY COMPONENT")
        lines.append("-" * 60)
        component_costs = bom.get_cost_by_component()
        for component, cost in sorted(component_costs.items(), key=lambda x: x[1], reverse=True):
            percentage = (cost / bom.total_monthly_cost) * 100
            lines.append(f"{component:<30} {bom.currency} {cost:>12,.2f} ({percentage:>5.1f}%)")
        lines.append("")
        
        # OS Breakdown
        lines.append("COST BREAKDOWN BY OS TYPE")
        lines.append("-" * 60)
        os_costs = bom.get_cost_by_os()
        for os_type, cost in sorted(os_costs.items(), key=lambda x: x[1], reverse=True):
            percentage = (cost / bom.total_monthly_cost) * 100
            lines.append(f"{os_type:<20} {bom.currency} {cost:>12,.2f} ({percentage:>5.1f}%)")
        lines.append("")
        
        # Detailed Line Items
        lines.append("DETAILED LINE ITEMS")
        lines.append("=" * 140)
        header_line = f"{'VM Name':<25} {'OS Type':<10} {'Component':<20} {'Description':<35} {'Qty':<10} {'Unit Price':<15} {'Total Cost':<15}"
        lines.append(header_line)
        lines.append("=" * 140)
        
        # Group by VM for better readability
        vm_items = {}
        for item in bom.line_items:
            if item.vm_name not in vm_items:
                vm_items[item.vm_name] = []
            vm_items[item.vm_name].append(item)
        
        # Sort VMs by total cost (descending)
        vm_totals = {vm_name: sum(item.total_cost for item in items) for vm_name, items in vm_items.items()}
        sorted_vms = sorted(vm_items.items(), key=lambda x: vm_totals[x[0]], reverse=True)
        
        for vm_name, items in sorted_vms:
            vm_total = sum(item.total_cost for item in items)
            
            for i, item in enumerate(items):
                # Show VM name and OS only on first line
                vm_display = vm_name[:24] if i == 0 else ""
                os_display = item.os_type.value if i == 0 else ""
                
                # Truncate description if too long
                description = item.description
                if len(description) > 34:
                    description = description[:31] + "..."
                
                # Format unit price and total cost with proper alignment
                unit_price_str = f"{bom.currency} {item.unit_price:,.4f}"
                total_cost_str = f"{bom.currency} {item.total_cost:,.2f}"
                
                line = f"{vm_display:<25} {os_display:<10} {item.component_type:<20} {description:<35} {item.quantity:<10.2f} {unit_price_str:<15} {total_cost_str:<15}"
                lines.append(line)
            
            # Add VM total and separator
            if len(items) > 1:
                vm_total_str = f"{bom.currency} {vm_total:,.2f}"
                lines.append(f"{'VM SUBTOTAL:':<25} {'':<10} {'':<20} {'':<35} {'':<10} {'':<15} {vm_total_str:<15}")
            
            lines.append("-" * 140)
        
        # Write to file
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        
        self.logger.info(f"Generated text BOM report: {output_file}")
        return output_file


class JSONReportGenerator(ReportGenerator):
    """JSON report generator for API integration"""
    
    @property
    def format_name(self) -> str:
        return "JSON"
    
    @property
    def file_extension(self) -> str:
        return "json"
    
    def generate_assessment_report(self, assessment: VMAssessment, output_file: str) -> str:
        """Generate JSON assessment report"""
        data = {
            "assessment_info": {
                "assessment_id": assessment.assessment_id,
                "assessment_name": assessment.assessment_name,
                "source_format": assessment.source_format,
                "assessment_date": assessment.assessment_date.isoformat(),
                "generated_at": datetime.now().isoformat()
            },
            "summary": self._serialize_stats(assessment.get_summary_stats()),
            "vms": []
        }
        
        for vm in assessment.vms:
            vm_data = {
                "vm_id": vm.vm_id,
                "vm_name": vm.vm_name,
                "os_type": vm.os_type.value,
                "os_config": vm.os_config,
                "cpu_cores": vm.cpu_cores,
                "memory_gb": vm.memory_gb,
                "total_storage_gb": vm.total_storage_gb,
                "power_state": vm.power_state.value,
                "cluster": vm.cluster,
                "host": vm.host,
                "datacenter": vm.datacenter,
                "annotation": vm.annotation,
                "storage": [
                    {
                        "capacity_gb": storage.capacity_gb,
                        "used_gb": storage.used_gb,
                        "storage_type": storage.storage_type.value,
                        "datastore": storage.datastore
                    }
                    for storage in vm.storage
                ],
                "networks": [
                    {
                        "interface_name": network.interface_name,
                        "mac_address": network.mac_address,
                        "ip_addresses": network.ip_addresses,
                        "network_name": network.network_name
                    }
                    for network in vm.networks
                ]
            }
            data["vms"].append(vm_data)
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        
        self.logger.info(f"Generated JSON assessment report: {output_file}")
        return output_file
    
    def generate_bom_report(self, bom: BillOfMaterials, output_file: str) -> str:
        """Generate JSON BOM report"""
        data = {
            "bom_info": {
                "assessment_id": bom.assessment_id,
                "currency": bom.currency,
                "pricing_source": bom.pricing_source,
                "pricing_date": bom.pricing_date.isoformat(),
                "generated_at": datetime.now().isoformat(),
                "notes": bom.notes
            },
            "summary": {
                "total_monthly_cost": bom.total_monthly_cost,
                "total_annual_cost": bom.total_annual_cost,
                "total_line_items": len(bom.line_items),
                "cost_by_component": bom.get_cost_by_component(),
                "cost_by_os": bom.get_cost_by_os()
            },
            "line_items": []
        }
        
        for item in bom.line_items:
            item_data = {
                "vm_id": item.vm_id,
                "vm_name": item.vm_name,
                "os_type": item.os_type.value,
                "component_type": item.component_type,
                "description": item.description,
                "quantity": item.quantity,
                "unit": item.unit,
                "unit_price": item.unit_price,
                "total_cost": item.total_cost,
                "monthly_cost": item.monthly_cost,
                "annual_cost": item.annual_cost,
                "pricing_model": item.pricing_model
            }
            data["line_items"].append(item_data)
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        
        self.logger.info(f"Generated JSON BOM report: {output_file}")
        return output_file
    
    def _serialize_stats(self, stats: Dict[str, Any]) -> Dict[str, Any]:
        """Convert datetime objects to ISO strings for JSON serialization"""
        serialized = {}
        for key, value in stats.items():
            if hasattr(value, 'isoformat'):  # datetime object
                serialized[key] = value.isoformat()
            else:
                serialized[key] = value
        return serialized